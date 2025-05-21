from os import path
import ezdxf
from openpyxl import Workbook, load_workbook
import datetime

def GetDXFData(DXFFile):

    Dir = path.dirname(DXFFile)

    doc = ezdxf.readfile(DXFFile)

    msp = doc.modelspace()
    Mleaders = msp.query("MULTILEADER")

    # CREATE EXCEL WORKBOOK
    wb = Workbook()
    ws = wb.active

    Row = 1
    Column = 1

    # READ DXF DATA AND WRITE TO EXCEL FILE
    for i in Mleaders:
        ws.cell(row = Row, column = Column).value = i.context.mtext.default_content
        ws.cell(row = Row, column = Column + 1).value = i.context.plane_origin[0]
        ws.cell(row = Row, column = Column + 2).value = i.context.plane_origin[1]
        
        Row += 1

    wb.save(Dir + "/" + str(datetime.datetime.now()).replace(":", "-") + ".xlsx")    