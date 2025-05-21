from os import path
from pyproj import Proj
from openpyxl import Workbook, load_workbook
import datetime
import xml.etree.ElementTree as ET
# from pykml.factory import KML_ElementMaker as KML
# from pykml import parser

import tkinter
import tkinter.filedialog

KMLFile = tkinter.filedialog.askopenfilename(title = 'Select KML file containing points', filetypes = [('KML files', '*.kml'), ('KMZ files', '*.kmz')])

Dir = path.dirname(KMLFile)

doc = ET.parse(KMLFile)

# root = doc.getroot()
# print(root.tag)

Projection = Proj(proj = 'utm', zone = 23, south = True, ellps = 'WGS84', preserve_units = False)

Points = []

x = 0
        
for pm in doc.iterfind('.//{*}Placemark'):
    Points.append([])
    Points[x].append(pm.find('{*}name').text)

    # print(pm.find('{*}name').text)
    
    for p in pm.iterfind('.//{*}Point'):
        Coordinates = p.find('{*}coordinates').text.split(',')
        X1 = Coordinates[0]
        Y1 = Coordinates[1]
        
        X2, Y2 = Projection(X1, Y1, inverse = False)
        
        Points[x].append(X2)
        Points[x].append(Y2)
        
        # print(X1, Y1)
        # print(p.find('{*}coordinates').text)
        
    x += 1
    
# CREATE EXCEL WORKBOOK
wb = Workbook()
ws = wb.active

Row = 1
Column = 1

# READ POINTS DATA AND WRITE TO EXCEL FILE
for p in Points:
    ws.cell(row = Row, column = Column).value = p[0]
    ws.cell(row = Row, column = Column + 1).value = p[1]
    ws.cell(row = Row, column = Column + 2).value = p[2]
    
    Row += 1

wb.save(Dir + "/" + str(datetime.datetime.now()).replace(":", "-") + ".xlsx")
        
# X2, Y2 = Projection(-40.6726325193965, -20.70470305630231, inverse = False)

# print(X2, Y2)

# print(Points)

# -40.6726325193965,-20.70470305630231
        
        
# nmsp = '{http://www.opengis.net/kml/2.2}'

# for pm in doc.iterfind('.//{0}Placemark'.format(nmsp)):
    # print(pm.find('{0}name'.format(nmsp)).text)