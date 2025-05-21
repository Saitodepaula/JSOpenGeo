from os import path
from openpyxl import Workbook, load_workbook
from pyproj import Proj
# from pyproj import CRS
# from pyproj.aoi import AreaOfInterest
# from pyproj.database import query_utm_crs_info
from pykml.factory import KML_ElementMaker as KML
from lxml import etree
import datetime

def MakeKML(Zone, isSouth, ExcelFile):

    Dir = path.dirname(ExcelFile)

    wb = load_workbook(filename = ExcelFile)
    ws = wb.active

    # Reads only cells with values
    Rows = len(tuple(ws.rows))

    p = Proj(proj = 'utm', zone = Zone, south = isSouth, ellps = 'WGS84', preserve_units = False)
    
    folder = KML.Folder()

    for row in range(1, Rows + 1):
        Name = ws.cell(row, 1).value
        X1 = ws.cell(row, 2).value
        Y1 = ws.cell(row, 3).value

        X2, Y2 = p (X1, Y1, inverse = True)
    
        coordinates = str(X2) + "," + str(Y2)
    
        NewPoint = KML.Placemark(
             KML.name(Name),
             KML.Point(
                KML.coordinates(coordinates)
                )
            )
            
        folder.append(NewPoint)
        
        print(Name)
        
    # convert the object tree into a string and write it into an output file
    with open(Dir + "/" + str(datetime.datetime.now()).replace(":", "-") + '.kml', 'wb') as file:
        file.write(etree.tostring(folder, pretty_print = True))
        
        
        
        
        
        
# ============================================================================================================
        
# TODO display map to find CRS or find by location (city name)

# utm_crs_list = query_utm_crs_info(
    # datum_name = "WGS 84",
    # area_of_interest = AreaOfInterest(
        # west_lon_degree = x,
        # south_lat_degree = y,
        # east_lon_degree = x,
        # north_lat_degree = y,
    # ),
# )
# utm_crs = CRS.from_epsg(utm_crs_list[0].code)

# print(utm_crs)