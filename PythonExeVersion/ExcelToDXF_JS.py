from os import path
import ezdxf
from ezdxf import zoom
from openpyxl import Workbook, load_workbook
import datetime

def GetExcelData(ExcelFile):

    Dir = path.dirname(ExcelFile)

    wb = load_workbook(filename = ExcelFile)
    ws = wb.active

    # Reads only cells with values
    Rows = len(tuple(ws.rows))

    # Create a new DXF R2010 drawing, official DXF version name: "AC1024"
    doc = ezdxf.new('R2010')

    # Add new entities to the modelspace:
    msp = doc.modelspace()

    # Create a block with the name 'POINTSYMBOL'
    PointSymbol = doc.blocks.new(name = 'POINTSYMBOL')
    PointSymbol.add_circle(ezdxf.math.Vec3(0, 0, 0), 2.0)
    PointSymbol.add_polyline2d([ezdxf.math.Vec3(-3, 0, 0), ezdxf.math.Vec3(3, 0, 0)])
    PointSymbol.add_polyline2d([ezdxf.math.Vec3(0, -3, 0), ezdxf.math.Vec3(0, 3, 0)])

    for row in range(1, Rows + 1):
    
        Name = ws.cell(row, 1).value
        X = ws.cell(row, 2).value
        Y = ws.cell(row, 3).value
        
        print(Name, X, Y)
        
        if Name == None:
            break
        
        P = ezdxf.math.Vec3(X, Y, 0)
             
        P1 = ezdxf.math.Vec3(X + 5, Y + 5, 0)
        
        msp.add_blockref('POINTSYMBOL', P)
        
        msp.add_circle(P, 2.0)    
        msp.add_leader([P, P1], 'EZDXF')
        
        ml_builder = msp.add_multileader_mtext()
        ml_builder.quick_leader(Name, target = ezdxf.math.Vec2(X, Y), segment1 = ezdxf.math.Vec2(5, 5))
        
        # print(Name, X, Y)
        
    zoom.extents(msp)

    doc.saveas(Dir + "/" + "PointsPlot" + str(datetime.datetime.now()).replace(":", "-") + ".dxf")